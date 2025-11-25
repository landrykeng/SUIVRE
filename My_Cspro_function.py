import sqlite3
import pandas as pd
import os
#import patoolib
import glob
from ftplib import FTP
import numpy as np
import zipfile
import streamlit as st
from value_sets_output import VALUE_SETS
from datetime import datetime
from openpyxl import load_workbook


#
dico_eleve_enseignant={
    "Etablissement":{101011: 'école publique de DARSO',
    102021: 'école publique de KANTALONG (PACARF)',
    102031: 'école publique de DARA EWA',
    103041: 'école publique de GASSOL-HARKA',
    103042: 'école publique de KOUI-LABBARE',
    104051: 'école publique bilingue de DHOHONG',
    104061: 'école publique de FADA',
    105071: 'école publiques de BEKA MATARI',
    105081: 'école publique de KOUBADJE',
    105091: 'école publique bilingue de DANG (PACARF)',
    201011: 'école publique de BOTOMO',
    201012: 'école publique de YAMBAYE',
    201033: 'école publique de BOTOMO',
    201034: 'école publique de YAMBAYE',
    201045: 'école publique de BOTOMO',
    201046: 'école publique de YAMBAYE',
    201021: 'école bilingue de GUENG',
    201022: 'école publique de ZAKAN',
    401011: 'école publique de BIR-PONDO',
    401012: 'école publique de MOKOLO 2',
    401013: 'école publique de NKOLBIKONG 3',
    401021: 'école publique de FRONTIERE',
    401022: 'école publique de NAGONDA',
    401023: 'école publique de SABAL VILLAGE',
    501031: 'école publique de SOURANDE',
    502041: 'école publique de MAKAMBARA (PIR)',
    502051: 'école publique de BOUBOUMA',
    502061: 'école publique de BOUNGOUR (PIR)',
    503071: 'école publique de NDOUDJIDA (PIR)',
    503081: 'école publique GOUNOUDA (PIR)',
    503091: 'école publique de BARABAYE',
    503101: 'école publique de GIDOUA',
    503111: 'école publique de VADA',
    503112: 'école publique de KAOUDOUS (PIR)',
    504121: 'école publique de KAYA (PIR)',
    504131: 'école publique de TCHADDE (PIR)',
    504141: 'école publique de MORDOK',
    504151: 'école publique de TOULOUM (Bilingue), MADIGALI (PIR)',
    505161: 'école publique de GOUDOUMBOUL (PIR)',
    505171: 'école publique bilingue de TOKOMBERE (PIR)',
    506181: 'école publique de DOUMBOULBAYE',
    506191: 'école publique de MALMAH (PIR)',
    506192: 'école publique de MARBA',
    506201: 'école publique de OUPAI KIRBI',
    601011: 'école publique Groupe 4 de Mbanga',
    601012: 'école Maternelle du quartier 9',
    601021: 'école annexe a Mbaressoumtou Carrière.',
    801011: 'école publique de Mbere',
    801012: 'école publique de Oumarou Ardo',
    801021: 'école publique de Maputki',
    801022: 'école publique de Laria',
    801031: 'école publique de Djalingo plateau',
    801041: 'école publique de Ouro-Barka',
    802051: 'école publique de More-Singai',
    802061: 'école publique de Bebere Gada-Mayo',
    901011: 'école primaire de MBVEH',
    902021: 'GS NWANGRI',
    1001011: 'école publique de BABETE G2',
    1001012: 'école publique de BATOUSSOP',
    1002021: 'école publique de KEKEM groupe 3',
    1003031: 'école publique de SUELAH',
    1003041: 'école publique de NGAZOM',
    1004051: 'école publique de FAMLENG',
    1005061: 'école publique de BANKOUOP',
    1005071: 'école publique de NAGHAM-NJIGOUMBE',
    1101011: "école publique d'ABELONG",
    1102021: "école publique Bilingue d'Efoulan"},
    "Région":{1: 'Adamaoua',
    2: 'Centre',
    3: 'Centre',
    4: 'Est',
    5: 'Extrême-Nord',
    6: 'Littoral',
    7: 'Littoral',
    8: 'Nord',
    9: 'Nord-Ouest',
    10: 'Ouest',
    11: 'Sud',
    12: 'Sud-Ouest'},
    "Departement":{ 10101: 'NGAOUNDAL',
    10202: 'KONTCHA',
    10203: 'TIGNERE',
    10304: 'BANYO',
    10405: 'DJOHONG',
    10406: 'MEIGANGA',
    10507: 'MARTAP',
    10508: 'NGANHA',
    10509: 'NGAOUNDERE 3ème',
    20101: 'BOKITO',
    20102: 'DEUK',
    20103: 'BOKITO',
    20104: 'BOKITO',
    40101: 'BERTOUA 1er',
    40102: 'GAROUA-BOULAI',
    50103: 'BOGO',
    50204: 'FOTOKOL',
    50205: 'GOULFEY',
    50206: 'MAKARY',
    50307: 'GOBO',
    50308: 'GUERE',
    50309: 'MAGA',
    50310: 'WINA',
    50311: 'YAGOUA',
    50412: 'DZIGUILAO',
    50413: 'KAELE',
    50414: 'MOUTOURWA',
    50415: 'TOULOUM',
    50516: 'KOLOFATA',
    50517: 'TOKOMBERE',
    50618: 'BOURRHA',
    50619: 'KOZA',
    50620: 'MOZOGO',
    60101: 'MBANGA',
    60102: 'NKONGSAMBA 1',
    80101: 'BARNDAKE',
    80102: 'BASCHEO',
    80103: 'GAROUA 3ème',
    80104: 'TOUROUA',
    80205: 'FIGUIL',
    80206: 'GUIDER',
    90101: 'KUMBO',
    90202: 'NKAMBE',
    100101: 'MBOUDA',
    100202: 'KEKEM',
    100303: 'PENKA-MICHEL',
    100304: 'SANTCHOU',
    100405: 'BAFOUSSAM 2ème',
    100506: 'FOUMBOT',
    100507: 'KOUOPTAMO',
    110101: 'DJOUM',
    110202: 'EFOULAN'},
    "Resultat":{1:"Rempli totalement",
                2:"Rempli partiellement",
                3:"Non rempli"}
        }

dico_maire={
    "Etablissement":dico_eleve_enseignant['Etablissement'],
    "Région": dico_eleve_enseignant['Région'],
    "Departement":dico_eleve_enseignant['Departement'],
    "Commune":{10101: "NGAOUNDAL",
    10202: "KONTCHA",
    10203: "TIGNERE",
    10304: "BANYO",
    10405: "DJOHONG",
    10406: "MEIGANGA",
    10507: "MARTAP",
    10508: "NGANHA",
    10509: "NGAOUNDERE 3ème",
    20101: "BOKITO",
    20102: "DEUK",
    40101: "BERTOUA 1er",
    40102: "GAROUA-BOULAI",
    50103: "BOGO",
    50204: "FOTOKOL",
    50205: "GOULFEY",
    50206: "MAKARY",
    50307: "GOBO",
    50308: "GUERE",
    50309: "MAGA",
    50310: "WINA",
    50311: "YAGOUA",
    50412: "DZIGUILAO",
    50413: "KAELE",
    50414: "MOUTOURWA",
    50415: "TOULOUM",
    50516: "KOLOFATA",
    50517: "TOKOMBERE",
    50618: "BOURRHA",
    50619: "KOZA",
    50620: "MOZOGO",
    60101: "MBANGA",
    60102: "NKONGSAMBA 1",
    80101: "BARNDAKE",
    80102: "BASCHEO",
    80103: "GAROUA 3ème",
    80104: "TOUROUA",
    80205: "FIGUIL",
    80206: "GUIDER",
    90101: "KUMBO",
    90202: "NKAMBE",
    100101: "MBOUDA",
    100202: "KEKEM",
    100303: "PENKA-MICHEL",
    100304: "SANTCHOU",
    100405: "BAFOUSSAM 2ème",
    100506: "FOUMBOT",
    100507: "KOUOPTAMO",
    110101: "DJOUM",
    110202: "EFOULAN"}}

dico_colonne_ens={
                #Enseignant
              "level-1-id":"ID",
              "s00q00":"Etablissement",
              "s00q01":"Région",
              "s00q02":"Département",
              "s00q03":"Arrondissement",
              "s00q04":"Localité",
              "s00q10":"Superviseur",
              "s00q11":"Controleur",
              "s00q12":"Enqueteur",
              "s00q13":"Date",
              "s00q17":"Résultat",
              "s00q17x":"Autre Résultat",
              "s00q17a":"Disponibilité",
              "g21a":"Longitude",
              "g21b":"Latitude",
              "hd":"Heure debut",
              "hf":"Heure fin",}


#Importation de la table depuis le fichier .csdb
def lire_csdb(csdb_path, table_name=None, columns=None, remove=False):
    conn = sqlite3.connect(csdb_path)
    cursor = conn.cursor()
    # Rename the table 'level-1' to 'level' if it exists
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='level-1';")
    if cursor.fetchone():
        cursor.execute("ALTER TABLE 'level-1' RENAME TO level;")
        conn.commit()
    # Trouver la table principale si non spécifiée
    if not table_name:
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
        table_name = cursor.fetchall()[0][0]  # première table
    
    # Construire la requête SQL
    if columns:
        columns_str = ", ".join(columns)
        query = f"SELECT {columns_str} FROM {table_name}"
        #query = f"SELECT * FROM {table_name}"
    else:
        query = f"SELECT * FROM {table_name}"
    
    # Lire les données
    other_file=csdb_path + ".lst"
    try:
        df = pd.read_sql_query(query, conn)
        conn.close()
        if remove==True:
            pass
            #os.remove(csdb_path)
            #os.remove(other_file)
        return df
    except Exception as e:
        print(f"Error reading {csdb_path}: {str(e)}")
        conn.close()
        return None
  

#annotation du dataframe pour la visualisation
def annoter_dataframe(df, dico_cspro):
    for col in df.columns:
        if col in dico_cspro :
            mapping = dico_cspro[col]
            try:
                df[col] = df[col].map(mapping)
            except Exception as e:
                print(f"Erreur d'annotation sur la variable {col} : {e}")
    return df

# Délécharger le fichier .csdb depuis un serveur FTP
import os
from ftplib import FTP, error_perm

def download_ftp_files():
    ftp_host = "217.112.80.251"
    ftp_path = "/CHEQUE2025/ENQUETE/DATA"
    ftp_user = "user_ins"
    ftp_pass = "123456@"

    try:
        ftp = FTP()
        ftp.connect(ftp_host, 21, timeout=60)
        ftp.login(ftp_user, ftp_pass)
        print("Connexion réussie")

        # 🔹 Se déplacer dans le bon dossier
        ftp.cwd(ftp_path)

        # 🔹 Récupérer la liste des sous-dossiers
        liste_brute = []
        ftp.retrlines("LIST", liste_brute.append)

        # 🔹 Extraire uniquement les noms des dossiers
        liste_dossiers = []
        for ligne in liste_brute:
            cols = ligne.split()
            nom = cols[-1]
            # Pour éviter les fichiers : on teste si c'est un dossier
            if ligne.upper().startswith("D") or " <DIR> " in ligne:
                liste_dossiers.append(nom)

        print("Dossiers trouvés :", liste_dossiers)

        # 🔹 Créer dossier local
        data_dir = "Data_Zip"
        os.makedirs(data_dir, exist_ok=True)

        # 🔹 Parcourir chaque dossier
        for dossier in liste_dossiers:
            print(f"\n📁 Dossier : {dossier}")

            try:
                ftp.cwd(f"{ftp_path}/{dossier}")
            except error_perm:
                print(f"❌ Impossible d'entrer dans : {dossier}")
                continue

            try:
                files = ftp.nlst()
            except:
                print("❌ Impossible de lister les fichiers")
                continue

            zip_files = [f for f in files if f.lower().endswith(".zip")]

            if not zip_files:
                print("⚠️ Aucun fichier .zip trouvé")
            else:
                for filename in zip_files:
                    local_filepath = os.path.join(data_dir, filename)
                    print(f"⬇️ Téléchargement : {filename}")

                    with open(local_filepath, "wb") as f:
                        ftp.retrbinary(f"RETR {filename}", f.write)

                    print(f"✔️ Fichier téléchargé : {local_filepath}")

            # Retourner dans le dossier principal
            ftp.cwd(ftp_path)

        ftp.quit()
        print("\n✔️ Téléchargement terminé avec succès.")

    except Exception as e:
        print("❌ Erreur :", e)


#Fonction d'extraction des données
def unzip_data_file(zip_filename):
    try:
        # Create extraction directory if it doesn't exist
        extract_dir = 'extracted_data'
        if not os.path.exists(extract_dir):
            os.makedirs(extract_dir)
            
        # Unzip the file
        with zipfile.ZipFile(zip_filename, 'r') as zip_ref:
            zip_ref.extractall(extract_dir)
            print(f"Successfully extracted {zip_filename} to {extract_dir}")
            
        # Return the list of extracted files
        extracted_files = os.listdir(extract_dir)
        return extracted_files
    
    except zipfile.BadZipFile:
        print(f"Error: {zip_filename} is not a valid zip file")
        return None
    except Exception as e:
        print(f"Error extracting {zip_filename}: {str(e)}")
        return None
      
#Fonction d'extraction entière
def Unzip_All_Files():
    zip_dir = 'Data_Zip'
    zip_files = [f for f in os.listdir(zip_dir) if f.endswith('.zip')]
    for zip_file in zip_files:
        zip_path = os.path.join(zip_dir, zip_file)
        extracted_files = unzip_data_file(zip_path)
        if extracted_files:  # Only delete if extraction was successful
            os.remove(zip_path) 
 

def extraire_et_agreger_csdb(rep_csdb):
    """
    Parcourt tous les fichiers .csdb du répertoire,
    extrait les tables cheque_rec, level, rejets, synthe,
    applique les merges, et renvoie deux dataframes agrégés :
    - df_rejets_final
    - df_synthe_final
    """

    liste_rejets = []
    liste_synthe = []
    liste_tabr=[]
    columns_label_rejet={
    "numfact":"Numéro de facture",
    "numch":"Numéro de chèque",
    'statinc':"Statut initial chèque",
    'noms':"Nom du répondant",
    'cat':"Catégorie",
    'typpres': "Type de prestation",
    'mfact':"Montant facturé",
    'mvaldmc':"Montant validé par le MC",
    'mvaldccm':"Montant validé par le Contr M",
    'mvaldcm':"Montant rejeté par le CM",
    'statch':"statut du chèque",
    'motif':"Motif de rejet_réhabilitation",
    'mfosa':"FOSA",
    'ms0q03':"NOM AIRE DE SANTE",
    's0q03':"Aire de santé",
    'ms0q02':"NOM DISTRICT",
    's0q01':"Région",
    's0q02':"District",
    's0q3':"Statut FOSA",
    's0q05':"Date collecte",
    'cenq':"Agent enqueteur", 
    'labmois':"Mois",  
}

    colums_label_synthese={
        'numf':"Numéro ligne",
        'type':"Type de prestation",
        "statut":"Prestation",
        'prest01':"TARIF UNITAIRE PRESTATION",
        'prest02':"NOMBRE TOTAL DE FACTURE RECU",
        'prest03':"MONTANT TOTAL FACTURE RECU",
        'prest04':"NOMBRE TOTAL DE FACTURE VALIDE MC",
        'prest05':"MONTANT TOTAL DE FACTURE NON VALIDE MC",
        'prest06':"MONTANT TOTAL DE FACTURE  VALIDE MC",
        'mfosa':"FOSA",
        's0q03':"Aire de santé",
        's0q01':"Région",
        's0q02':"District",
        's0q3':"Statut FOSA",
        's0q05':"Date collecte",
        'ms0q03':"NOM AIRE DE SANTE",
        'ms0q02':"NOM DISTRICT",
        'cenq':"Agent enqueteur",
        'labmois':"Mois",
    }
    
    colums_label_tabr={
        'num':"Nombre de chèque à saisir",
        'mfosa':"FOSA",
        's0q03':"Aire de santé",
        's0q01':"Région",
        's0q02':"District",
        's0q3':"Statut FOSA",
        's0q05':"Date collecte",
        'cenq':"Agent enqueteur",
        'labmois':"Mois",
        'ms0q03':"NOM AIRE DE SANTE",
        'ms0q02':"NOM DISTRICT",
    }
    # Parcours de tous les fichiers .csdb
    for f in os.listdir(rep_csdb):
        if f.lower().endswith(".csdb"):
            chemin = os.path.join(rep_csdb, f)
            print(f"📂 Traitement de : {f}")

            try:
                # Lecture des tables
                cheque = lire_csdb(chemin, table_name="cheque_rec")
                lev = lire_csdb(chemin, table_name="level")
                rejet = lire_csdb(chemin, table_name="rejets")
                synthe = lire_csdb(chemin, table_name="synthe")
                tabr=lire_csdb(chemin, table_name="tabr")

                
                tabr=tabr.merge(cheque, on="level-1-id", how="left")
                liste_tabr.append(tabr)
                # Merge cheque + level
                cheque = cheque.merge(lev, on="level-1-id", how="left")

                # Merge rejet + cheque
                rejet = rejet.merge(cheque, on="level-1-id", how="left")
                liste_rejets.append(rejet)

                # Merge synthe + cheque
                synthe = synthe.merge(cheque, on="level-1-id", how="left")
                liste_synthe.append(synthe)

            except Exception as e:
                print(f"❌ Erreur sur le fichier {f} : {e}")
                continue

    # Agrégation finale
    if liste_rejets:
        df_rejets_final = pd.concat(liste_rejets, ignore_index=True)
    else:
        df_rejets_final = pd.DataFrame()

    if liste_synthe:
        df_synthe_final = pd.concat(liste_synthe, ignore_index=True)
    else:
        df_synthe_final = pd.DataFrame()
    if liste_tabr:
        tabr_final=pd.concat(liste_tabr, ignore_index=True)
    else:
        tabr_final=pd.DataFrame()
        
    
    # Sélectionner toutes les colonnes numériques
    #num_cols_rejet = df_rejets_final.select_dtypes(include=["float", "int"]).columns
    #df_rejets_final[num_cols_rejet]=df_rejets_final[num_cols_rejet].astype("Int64")
    
    #num_cols_synthese = df_synthe_final.select_dtypes(include=["float", "int"]).columns
    #df_synthe_final[num_cols_synthese]=df_synthe_final[num_cols_synthese].astype("Int64")
    
    #Maping
    df_rejets_final['statinc']=df_rejets_final['statinc'].replace(VALUE_SETS['statinc'])
    df_rejets_final['cat']=df_rejets_final['cat'].replace(VALUE_SETS['cat'])
    df_rejets_final['typpres']=df_rejets_final['typpres'].replace(VALUE_SETS['typpres'])
    df_rejets_final['statch']=df_rejets_final['statch'].replace(VALUE_SETS['statch'])
    df_rejets_final['motif']=df_rejets_final['motif'].replace(VALUE_SETS['motif'])
    df_rejets_final['s0q03']=df_rejets_final['s0q03'].replace(VALUE_SETS['s0q03'])
    df_rejets_final['s0q01']=df_rejets_final['s0q01'].replace(VALUE_SETS['s0q01'])
    df_rejets_final['s0q02']=df_rejets_final['s0q02'].replace(VALUE_SETS['s0q02'])
    df_rejets_final['s0q3']=df_rejets_final['s0q3'].replace(VALUE_SETS['s0q3'])
    #df_rejets_final['mois']=df_rejets_final['mois'].replace(VALUE_SETS['mois'])
    
    df_synthe_final['type']=df_synthe_final['type'].replace(VALUE_SETS['type'])
    df_synthe_final['statut']=df_synthe_final['statut'].replace(VALUE_SETS['typpres'])
    df_synthe_final['s0q03']=df_synthe_final['s0q03'].replace(VALUE_SETS['s0q03'])
    df_synthe_final['s0q01']=df_synthe_final['s0q01'].replace(VALUE_SETS['s0q01'])
    df_synthe_final['s0q02']=df_synthe_final['s0q02'].replace(VALUE_SETS['s0q02'])
    df_synthe_final['s0q3']=df_synthe_final['s0q3'].replace(VALUE_SETS['s0q3'])

    tabr_final['s0q03']=tabr_final['s0q03'].replace(VALUE_SETS['s0q03'])
    tabr_final['s0q01']=tabr_final['s0q01'].replace(VALUE_SETS['s0q01'])
    tabr_final['s0q02']=tabr_final['s0q02'].replace(VALUE_SETS['s0q02'])
    tabr_final['s0q3']=tabr_final['s0q3'].replace(VALUE_SETS['s0q3'])
        
    #Rétension des colonne utiles et renomage
    df_rejets_final = df_rejets_final.drop(columns=["cheque_rec-id", "case-id", "numfact_1","mois","action1"])
    df_rejets_final=df_rejets_final.rename(columns=columns_label_rejet)
    
    df_synthe_final = df_synthe_final.drop(columns=["synthe-id", "action", "cheque_rec-id","case-id","nlg","mois"])
    df_synthe_final=df_synthe_final.rename(columns=colums_label_synthese)
    
    tabr_final = tabr_final.drop(columns=["tabr-id", "ftab1", "cheque_rec-id","nlg"])
    tabr_final=tabr_final.rename(columns=colums_label_tabr)
    
    
    fichier_data="Data_collected.xlsx"
    
    
    
    with pd.ExcelWriter(fichier_data, engine="xlsxwriter") as writer:
        df_rejets_final.to_excel(writer, sheet_name="Rejet", index=False)
        df_synthe_final.to_excel(writer, sheet_name="Synthese", index=False)
        tabr_final.to_excel(writer, sheet_name="Initial", index=False)
    
    wb = load_workbook(fichier_data)

    # Sélectionner la feuille (la créer si elle n'existe pas)
    if "Update" not in wb.sheetnames:
        wb.create_sheet("Update")
    ws = wb["Update"]   # mettre le nom exact de la feuille

    # Écrire dans une cellule
    ws["A1"] = datetime.now()

    # Sauvegarder
    wb.save(fichier_data)
    return df_rejets_final, df_synthe_final, tabr_final



